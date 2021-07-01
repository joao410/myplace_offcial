from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.forms import modelformset_factory
import csv
from .models import  User
from helpdesk.models import  Chamado, Image , ImageLink, Chat
from .models import   UsuarioCorporativo, UsuarioEndereco, UsuarioTrabalho,UsuarioDocumentos, Empresa, ImagePerfil, UsuarioPessoal,Cargo,Contabancaria
from helpdesk.forms import ImageForm, ImageForms
from django.http import HttpResponseRedirect
from django.contrib import messages
from helpdesk.filters import ChamadoFilter
from datetime import date, datetime, timedelta
from django.contrib.auth.models import User, Group, GroupManager
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.forms import modelformset_factory
from django.http import JsonResponse, HttpResponse
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import load_workbook
from shutil import copyfile
import pandas as pd
from decimal import Decimal
import xlwt



# Create your views here.

@login_required(login_url='/authentication/login')    
def add_usuarios(request):
    user = request.user
    admin = Chamado.objects.all()        
    usuarioC = UsuarioCorporativo.objects.get(usuario=user)
    g = Group.objects.all()
    grupo= usuarioC.grupo
    chamados = Chamado.objects.filter(grupo=grupo).order_by("-id")    
    chamados_abertos = Chamado.objects.filter(active=True,grupo=grupo).order_by("-id")
    grupos= usuarioC.grupo.name
    codigo = usuarioC.codigo.nome   
    imageP = ImagePerfil.objects.get(nome= codigo)
    if not imageP.image:
      imageP = ImagePerfil.objects.get(nome= "padrao")
    empresa = Empresa.objects.all()
    cargos = Cargo.objects.all().order_by("cargo")
    filtro = ChamadoFilter()
    context = {
    'chamados': chamados,
    'chamados_abertos': chamados_abertos,
    'grupos': grupos,
    'usuarioC':usuarioC,
    'user' : user,
    'g':g,  
    'grupo':grupo,
    'empresa':empresa,
    'admin':admin, 
    'filtro': filtro,
    'imageP':imageP, 
    'cargos':cargos,  
    'form':ImageForm,
            }  
    if UsuarioPessoal.objects.all():
            Usuario_id = UsuarioPessoal.objects.all().order_by('-id')[0].id
                

            codigo = Usuario_id + 10001
    else:
                    codigo = 10001
                                
    if request.method == 'POST' and 'add_usu' in request.POST:  
        
        #pessoal#
        try :
            nome = request.POST["nome"]
            nome = nome.upper()
            apelido = request.POST["apelido"]
            cpf = request.POST["cpf"]
            genero = request.POST["genero"]
            cor = request.POST["cor"]
            ecivil = request.POST["ecivil"]
            escolaridade = request.POST["escolaridade"]
            datanasci = request.POST["datanasci"]
            municipionasc = request.POST["municipionasc"]
            ufnasci = request.POST["ufnasci"]
            paisnasci = request.POST["paisnasci"]
            nasciona = request.POST["nasciona"]
            mae = request.POST["mae"]
            pai = request.POST["pai"]
            pis = request.POST["pis"]
            teleitor =  request.POST["titulo"]
            ctrabalho = request.POST["carteira"]       
            serie = request.POST["serie"]
            uf = request.POST["ufc"]
            emissao = request.POST["emissao"]
            celp = request.POST["celp"]
        except:
            pass    
        #trabalho# 
        try:
            emp = request.POST["empresa"]
            dep = request.POST["departamento"]       
            cargo = request.POST["cargo"]
            vtrans = request.POST["vtrans"]
            admissao = request.POST["admissao"]
            demissao = request.POST["demissao"]
            tipo = request.POST["tipoadmissao"]
            indica =  request.POST["indica"]
            priempr = request.POST["priempr"]
            rtrab = request.POST["rtrab"]
            rprev = request.POST["rprev"]
            rjorn = request.POST["rjorn"]
            naativ = request.POST["naativ"]
            cat = request.POST["cat"]
            codf = request.POST["codf"]
            carh = request.POST["carh"]
            unisa = request.POST["unisa"]
            obs = request.POST["obs"]
            salvari = request.POST["salvari"]
            if not salvari:
                salvari =0.00
        except:
            pass
        #conta#
        try:
            banco = request.POST["banco"]
            agencia = request.POST["agencia"]
            conta = request.POST["conta"]
        except:
            pass    
        #documento#
        try:
            documento = request.POST["documento"]
            ndocumento = request.POST["numero"]
            oe = request.POST["oe"]
            de = request.POST["expedicao"]
            validade = request.POST["validade"]
        except:
            pass     
        #endereço#
        try:
            cep = request.POST["cep"]
            tipo = request.POST["tipo"]
            num = request.POST["num"]
            ufatual = request.POST["ufatual"]
            muniatual = request.POST["muniatual"]
            bairro = request.POST["bairro"]
            logradouro = request.POST["logradouro"]
            complemento = request.POST["complemeno"]
            pais = request.POST["pais"]
        except:
            pass
        #corporativo#
        try:
            email1 = request.POST["email1"]       
            email2 = request.POST["email2"]       
            skype = request.POST["skype"]       
            cel = request.POST["cel"]       
            tel = request.POST["tel"]       
            ramal = request.POST["ramal"]       
            usuario = request.POST["usuario"] 
            usuario = usuario.upper()      
            senha = request.POST["senha"]  
            senha=senha.upper() 
            repassword = request.POST["senha"]    
            repassword = repassword.upper()
            
            grupo = request.POST["grupo"] 
            gs = Group.objects.get(name=grupo)   
        except:
            pass
        if vtrans == "SIM":
           vtrans =  "VALE TRANSPORTE"
        if vtrans == "NÃo":
           vtrans =  "AJUDA DE CUSTO"  
        if not vtrans:  
           vtrans = "---"            
    
                  
        if not User.objects.filter(username=usuario).exists():
                        
            if len(senha) < 6:
                messages.error(request, "Senha muito curta(<6)")
                return render(request, 'users/add_usuarios.html', context)
            
            if senha != repassword:
                messages.error(request, "As senhas nao batem")
                return render(request, 'users/add_usuarios.html', context)
            user = User.objects.create(username=usuario,email=email1,first_name=usuario)
            user.set_password(senha)
            user.is_active = True
            user.save()
            users = User.objects.get(username=usuario)
            try:
                UsuarioPessoal.objects.create(codigo=codigo,nome=nome,genero=genero,apelido=apelido,celpessoal=celp,cpf=cpf,cor=cor,ecivil=ecivil,escolaridade=escolaridade,pis=pis,tituloeleitor=teleitor,carteiratrabalho=ctrabalho,serie=serie,ufcarteiratrabalho=uf,datacarteiratrabalho=emissao,datanacimento=datanasci,ufnacimento=ufnasci,municipionacimento=municipionasc,paisnacimento=paisnasci,paisnacionalidade=nasciona,nomemae=mae,nomepai=pai)
            except:
                UsuarioPessoal.objects.create(codigo=codigo,nome=nome,genero=genero,apelido=apelido,cpf=cpf,celpessoal=celp,cor=cor,ecivil=ecivil,escolaridade=escolaridade,pis=pis,tituloeleitor=teleitor,carteiratrabalho=ctrabalho,serie=serie,ufcarteiratrabalho=uf,datanacimento=datanasci,ufnacimento=ufnasci,municipionacimento=municipionasc,paisnacimento=paisnasci,paisnacionalidade=nasciona,nomemae=mae,nomepai=pai)
            u = UsuarioPessoal.objects.get(nome=nome)
            c= Cargo.objects.get(cargo=cargo)
            try:
                usua = UsuarioTrabalho.objects.create(codigo=u,departamento=dep,empresa=emp,cargo=c,valetransporte=vtrans,dataadmissao=admissao,datademissao=demissao,indicativoadmissao=indica,primeiroemprego=priempr,regimetrabalho=rtrab,regimeprevidenciario=rprev,regimejornada=rjorn,naturezaatividade=naativ,categoria=cat,codigofuncao=codf,cargahorariam=carh,unidadesalarial=unisa,salariovariavel=salvari,obs=obs)
                t = UsuarioTrabalho.objects.get(codigo=u)
            except:
                pass    
            try:
                do = UsuarioDocumentos.objects.create(codigo=u,documento=documento,numerodocumento=ndocumento,orgao=oe,dataexpedissao=de,validade=validade)
                d =  UsuarioDocumentos.objects.get(codigo=u)
            except:
                pass    
            try:
                en = UsuarioEndereco.objects.create(codigo=u, cep=cep,tipo=tipo,logradouro=logradouro,numero=num,ufatual=ufatual,municipioatul=muniatual,bairroatual=bairro,complemento=complemento,pais=pais)
                e = UsuarioEndereco.objects.get(codigo=u)
            except:
                pass    
            try:
                co = Contabancaria.objects.create(codigo=u,banco=banco,conta=conta,agencia=agencia)
            except:
                pass
            try:    
                usuar  = UsuarioCorporativo.objects.create(codigo=u,trabalho=t,documento=d,endereco=e,email=email1,emailCorporativo=email2,skype=skype,telefone=cel,tel=tel,ramal=ramal,usuario=users,grupo=gs,banco=co)
            except:
                 usuar  = UsuarioCorporativo.objects.create(codigo=u,email=email1,emailCorporativo=email2,skype=skype,telefone=cel,tel=tel,ramal=ramal,usuario=users,grupo=gs)
            
            form = ImageForm(request.POST, request.FILES)
            if form.is_valid():
                nome = request.POST['usuario']
                img = form.cleaned_data.get("imagem") 
                obs=''
                users = User.objects.get(username=usuario)
                usuari = UsuarioCorporativo.objects.get(usuario=users)
                us = usuari.codigo.nome

                        
                obj =ImagePerfil.objects.create(image=img,obs=obs,nome=us)
                obj.save()
                messages.success(request, 'Imagem adicionada')
            else:
                messages.error(request, 'Imagem não adicionada')
            messages.success(request, "Usuario criado com sucesso")
            return redirect( 'presidente')
        elif User.objects.filter(username=usuario).exists and UsuarioCorporativo.objects.filter(usuario=user).exists:
              messages.error(request, 'Usuario existente')
        else:
                users = User.objects.get(username=usuario)
                usu = Usuarios.objects.create(name=nome,empresa=emp, departamento=dep,cargo=cargo,email=email1,email_corporativo=email2,skype=skype,telefone=cel,tel=tel,ramal=ramal,usuario=users,password=senha,grupo=gs)
                usu.save()
                form = ImageForm(request.POST, request.FILES)
                if form.is_valid():
                    nome = request.POST['usuario']
                    img = form.cleaned_data.get("imagem") 
                    obs=''
                    users = User.objects.get(username=usuario)
                    usuari = Usuarios.objects.get(usuario=users)
                    obj =ImagePerfil.objects.create(usuario=usuari,image=img,obs=obs,nome=nome)
                    obj.save()
                    messages.success(request, 'Imagem adicionada')

                
    return render(request, 'users/add_usuarios.html', context)        
    
    
@login_required(login_url='/authentication/login')    
def edit_usuarios(request,id):
    user = request.user
    admin = Chamado.objects.all()        
    usuarioC = UsuarioCorporativo.objects.get(usuario=user)
    g = Group.objects.all()
    grupo= usuarioC.grupo
   
    chamados = Chamado.objects.filter(grupo=grupo).order_by("-id")    
    chamados_abertos = Chamado.objects.filter(active=True,grupo=grupo).order_by("-id")
    grupos= usuarioC.grupo.name
    codigo = usuarioC.codigo.nome
    empresa = Empresa.objects.all()
    cargos = Cargo.objects.all().order_by("cargo")
    filtro = ChamadoFilter()
    usu = UsuarioCorporativo.objects.get(pk=id)
    usup = UsuarioPessoal.objects.get(codigo= usu.codigo.codigo)
    usut = UsuarioTrabalho.objects.get(codigo=usu.codigo)
    usud = UsuarioDocumentos.objects.get(codigo=usu.codigo)
    usue = UsuarioEndereco.objects.get(codigo=usu.codigo)
    cod =usu.codigo.nome

    
    imageP = ImagePerfil.objects.get(nome= codigo)
    if not imageP.image:
      imageP = ImagePerfil.objects.get(nome= "padrao")
    image = ImagePerfil.objects.get(nome=cod)
    if not image.image:
      image = ImagePerfil.objects.get(nome= "padrao")

    if request.method == 'POST' and 'next' in request.POST:
        id = id + 1
        try:
            usu = UsuarioCorporativo.objects.get(pk=id)
            usup = UsuarioPessoal.objects.get(codigo= usu.codigo.codigo)
            usut = UsuarioTrabalho.objects.get(codigo=usu.codigo)
            usud = UsuarioDocumentos.objects.get(codigo=usu.codigo)
            usue = UsuarioEndereco.objects.get(codigo=usu.codigo)
            cod =usu.codigo.nome
            imageP = ImagePerfil.objects.get(nome= codigo)
            if not imageP.image:
                imageP = ImagePerfil.objects.get(nome= "padrao")
            image = ImagePerfil.objects.get(nome=cod)
            if not image.image:
                image = ImagePerfil.objects.get(nome= "padrao")
        
        except:
            
            try:
                us = UsuarioCorporativo.objects.all()
                indice= len(us)
                i=us[indice].id
                while id > 0 and id < i:
                    try:
                        id = id+1  
                        usu = UsuarioCorporativo.objects.get(pk=id)
                        usup = UsuarioPessoal.objects.get(codigo= usu.codigo.codigo)
                        usut = UsuarioTrabalho.objects.get(codigo=usu.codigo)
                        usud = UsuarioDocumentos.objects.get(codigo=usu.codigo)
                        usue = UsuarioEndereco.objects.get(codigo=usu.codigo)
                        cod =usu.codigo.nome
                        imageP = ImagePerfil.objects.get(nome= codigo)
                        if not imageP.image:
                            imageP = ImagePerfil.objects.get(nome= "padrao")
                        image = ImagePerfil.objects.get(nome=cod)
                        if not image.image:
                            image = ImagePerfil.objects.get(nome= "padrao")
                        break   
                    except:
                        pass 
            except:
                pass
    if request.method == 'POST' and 'previos' in request.POST:
        id = id - 1
        try:
            usu = UsuarioCorporativo.objects.get(pk=id)
            usup = UsuarioPessoal.objects.get(codigo= usu.codigo.codigo)
            usut = UsuarioTrabalho.objects.get(codigo=usu.codigo)
            usud = UsuarioDocumentos.objects.get(codigo=usu.codigo)
            usue = UsuarioEndereco.objects.get(codigo=usu.codigo)
            cod =usu.codigo.nome
            imageP = ImagePerfil.objects.get(nome= codigo)
            if not imageP.image:
                imageP = ImagePerfil.objects.get(nome= "padrao")
            image = ImagePerfil.objects.get(nome=cod)
            if not image.image:
                image = ImagePerfil.objects.get(nome= "padrao")
        except:
           while id >0: 
                id = id-1  
                try:
                    usu = UsuarioCorporativo.objects.get(pk=id)
                    usup = UsuarioPessoal.objects.get(codigo= usu.codigo.codigo)
                    usut = UsuarioTrabalho.objects.get(codigo=usu.codigo)
                    usud = UsuarioDocumentos.objects.get(codigo=usu.codigo)
                    usue = UsuarioEndereco.objects.get(codigo=usu.codigo)
                    cod =usu.codigo.nome
                    imageP = ImagePerfil.objects.get(nome= codigo)
                    if not imageP.image:
                        imageP = ImagePerfil.objects.get(nome= "padrao")
                    image = ImagePerfil.objects.get(nome=cod)
                    if not image.image:
                        image = ImagePerfil.objects.get(nome= "padrao")
                except:
                    pass
                       
             
                   
                    
               


    context = {
        'chamados': chamados,
        'chamados_abertos': chamados_abertos,
        'grupos': grupos,
        'usuarioC':usuarioC,
        'user' : user,
        'g':g,  
        'grupo':grupo,
        'admin':admin, 
        'filtro': filtro,
        'imageP':imageP,   
        'form':ImageForm,
        'usu': usu,
        'image':image,
        'empresa':empresa,
        'cargos':cargos, 
            }  
    if request.method == 'POST' and 'edit_usu' in request.POST:  
        nome = request.POST["nome"]
        apelido = request.POST["apelido"]
        cpf = request.POST["cpf"]
        genero = request.POST["genero"]
        cor = request.POST["cor"]
        ecivil = request.POST["ecivil"]
        escolaridade = request.POST["escolaridade"]
        datanasci = request.POST["datanasci"]
        municipionasc = request.POST["municipionasc"]
        ufnasci = request.POST["ufnasci"]
        paisnasci = request.POST["paisnasci"]
        nasciona = request.POST["nasciona"]
        mae = request.POST["mae"]
        pai = request.POST["pai"]
        pis = request.POST["pis"]
        teleitor =  request.POST["titulo"]
        ctrabalho = request.POST["carteira"]       
        serie = request.POST["serie"]
        uf = request.POST["ufc"]
        emissao = request.POST["emissao"]
        celp = request.POST["celp"]
        #trabalho# 
        
        emp = request.POST["empresa"]
        dep = request.POST["departamento"]       
        cargo = request.POST["cargo"]
        vtrans = request.POST["vtrans"]
        admissao = request.POST["admissao"]
        demissao = request.POST["demissao"]
        tipo = request.POST["tipoadmissao"]
        indica =  request.POST["indica"]
        priempr = request.POST["priempr"]
        rtrab = request.POST["rtrab"]
        rprev = request.POST["rprev"]
        rjorn = request.POST["rjorn"]
        naativ = request.POST["naativ"]
        cat = request.POST["cat"]
        codf = request.POST["codf"]
        carh = request.POST["carh"]
        unisa = request.POST["unisa"]
        obs = request.POST["obs"]
        salvari = request.POST["salvari"]
        #conta#
        banco = request.POST["banco"]
        agencia = request.POST["agencia"]
        conta = request.POST["conta"]
        #documento#
        documento = request.POST["documento"]
        ndocumento = request.POST["numero"]
        oe = request.POST["oe"]
        de = request.POST["expedicao"]
        validade = request.POST["validade"]
        #endereço#
        cep = request.POST["cep"]
        tipoe = request.POST["tipo"]
        num = request.POST["num"]
        ufatual = request.POST["ufatual"]
        muniatual = request.POST["muniatual"]
        bairro = request.POST["bairro"]
        logradouro = request.POST["logradouro"]
        complemento = request.POST["complemeno"]
        pais = request.POST["pais"]
        #corporativo#
        email1 = request.POST["email1"]       
        email2 = request.POST["email2"]       
        skype = request.POST["skype"]       
        cel = request.POST["cel"]       
        tel = request.POST["tel"]       
        ramal = request.POST["ramal"]       
        usuario = request.POST["usuario"]       
        
        grupo = request.POST["grupo"] 
  
        gs = Group.objects.get(name=grupo)  
        
        if vtrans == "SIM":
           vtrans =  "VALE TRANSPORTE"
        if vtrans == "NÃo":
           vtrans =  "AJUDA DE CUSTO"  
        if not vtrans:  
             vtrans = "---"     

        if  User.objects.filter(username=usuario).exists():  
                   
            
            user = User.objects.get(username=usuario)
            user.is_active = True
            user.save()
            users = User.objects.get(username=usuario)
            usup.nome = nome
            usup.apelido= apelido
            usup.cpf = cpf
            usup.pis = pis
            usup.tituloeleitor = teleitor
            usup.carteiratrabalho = ctrabalho
            usup.serie = serie
            usup.ufcarteiratrabalho = uf
            try:
                usup.datacarteiratrabalho = datetime.strptime(de,'%d/%m/%Y')
            except:
                usup.datacarteiratrabalho = None      
            usup.genero = genero
            usup.cor = cor
            usup.ecivil= ecivil
            usup.celpessoal= celp
            usup.escolaridade  = escolaridade
            usup.datanacimento = datetime.strptime(datanasci ,'%d/%m/%Y')
            usup.ufnacimento = ufnasci
            usup.municipionacimento = municipionasc
            usup.paisnacimento = paisnasci
            usup.paisnacionalidade = nasciona
            usup.nomemae = mae
            usup.nomepai = pai
            usup.save()
            if Contabancaria.objects.filter(codigo=usup.codigo).exists():  
                conta = Contabancaria.objects.get(codigo=usup.codigo)
                conta.banco = banco
                conta.agencia = agencia
                conta.conta = conta
                conta.save()
            else:
                u=UsuarioPessoal.objects.get(pk=usup.id)
                conta = Contabancaria.objects.create(codigo=u,banco=banco,agencia=agencia,conta=conta)
                conta.save()
            usut.empresa = emp
            usut.departamento = dep
            usut.cargo = Cargo.objects.get(cargo=cargo)
            usut.valetransporte = vtrans
            try:
                usut.dataadmissao = datetime.strptime(admissao,'%d/%m/%Y')
            except:
                usut.dataadmissao = None
            try:   
                usut.datademissao = datetime.strptime(demissao,'%d/%m/%Y')
            except:
                usut.datademissao = None
            usut.tipoAdmissao = tipo
            usut.indicativoadmissao= indica
            usut.primeiroemprego = priempr
            usut.regimetrabalho = rtrab
            usut.regimeprevidenciario = rprev
            usut.regimejornada = rjorn
            usut.naturezaatividade = naativ
            usut.categoria = cat
            usut.codigofuncao = codf
            usut.cargahorariam = carh
            usut.unidadesalarial = unisa
            try:
                usut.salariovariavel = Decimal(salvari.replace(',','.'))
            except :
                usut.salariovariavel = 0
            usut.obs = obs
            usut.save()

            usud.documento= documento
            usud.numerodocumento = ndocumento
            usud.orgao = oe
            try:
                usud.dataexpedissao =  datetime.strptime(de,'%d/%m/%Y')
            except:
               usud.dataexpedissao = None
            try:
                usud.validade =  datetime.strptime(validade,'%d/%m/%Y')
            except:
                usud.validade = None
            usud.save()
            usue.cep = cep
            usue.tipo = tipoe
            usue.logradouro = logradouro
            usue.numero = num
            usue.ufatual = ufatual
            usue.MunicipioAtul = muniatual
            usue.bairroatual = bairro
            usue.complemento = complemento
            usue.pais = pais
            usue.save()
            usu.email = email1
            usu.emailCorporativo = email2
            usu.skype = skype
            usu.telefone = cel
            usu.tel = tel
            usu.ramal = ramal
            usu.usuario = users
            usu.grupo= gs
            usu.save()
            image = ImagePerfil.objects.get(nome = cod)
            image.nome = usup.nome
            image.save() 
            messages.success(request, "Usuario editado  com sucesso")

    



    if request.method == 'POST' and 'exc_usu' in request.POST: 
        id = request.POST['exc_usu']
        usuario= Usuarios.objects.get(pk=id)
        usuario.delete()
        return redirect( 'presidente')         
                
    return render(request, 'users/edit_usuarios.html', context)        
@login_required(login_url='/authentication/login')    
def perfis(request):
    ids = request.POST["perfil"]
    user = request.user
    admin = Chamado.objects.all()        
    usuarioC = UsuarioCorporativo.objects.get(usuario=user)
    usuario = UsuarioCorporativo.objects.get(pk=ids)
    grupo= usuarioC.grupo
    codigo = usuarioC.codigo
    chamados = Chamado.objects.filter(grupo=grupo).order_by("-id")    
    chamados_abertos = Chamado.objects.filter(active=True,grupo=grupo).order_by("-id")
    grupos= usuarioC.grupo.name
    imageP = ImagePerfil.objects.get(nome= codigo.nome)
    if not imageP.image:
      imageP = ImagePerfil.objects.get(nome= "padrao")
    img = ImagePerfil.objects.get(nome=usuario.codigo.nome)
    if not img.image:
      img = ImagePerfil.objects.get(nome= "padrao")
    filtro = ChamadoFilter()
    context = {
    'chamados': chamados,
    'chamados_abertos': chamados_abertos,
    'grupos': grupos,
    'usuarioC':usuarioC,
    'user' : user,  
    'grupo':grupo,
    'admin':admin, 
    'filtro': filtro,
    'imageP':imageP,   
    'img':img, 
    'usuario':usuario,
            }  
    return render(request, 'users/perfis.html', context)        
    
@login_required(login_url='/authentication/login')    
def perfisuser(request):
    ids = request.POST["perfil"]
    user = request.user
    admin = Chamado.objects.all()        
    usuarioC = UsuarioCorporativo.objects.get(usuario=user)
    usuario = UsuarioCorporativo.objects.get(pk=ids)
    grupo= usuarioC.grupo
    chamados = Chamado.objects.filter(grupo=grupo).order_by("-id")    
    chamados_abertos = Chamado.objects.filter(active=True,grupo=grupo).order_by("-id")
    grupos= usuarioC.grupo.name
    codigo = usuarioC.codigo.nome
    cod =usuario.codigo.nome
    
    try:
        image = ImagePerfil.objects.get(nome=cod)
        if not image.image:
            image = ImagePerfil.objects.get(nome="padrao")
    except:
        image = ImagePerfil.objects.get(nome= "padrao") 
    imageP = ImagePerfil.objects.get(nome=codigo) 
    try:
        imageP = ImagePerfil.objects.get(nome= codigo.nome)
        if not imageP.image:
            imageP = ImagePerfil.objects.get(nome= "padrao")
    except:
        imageP = ImagePerfil.objects.get(nome= "padrao")
    filtro = ChamadoFilter()
    context = {
    'chamados': chamados,
    'chamados_abertos': chamados_abertos,
    'grupos': grupos,
    'usuarioC':usuarioC,
    'user' : user,  
    'grupo':grupo,
    'admin':admin, 
    'filtro': filtro,
    'imageP':imageP,
    'image':image, 
    'usuario':usuario,
            }  
   
        
               
    return render(request, 'users/perfisuser.html', context)      
@login_required(login_url='/authentication/login')    
def home(request):
    user = request.user
    admin = Chamado.objects.all()
          
    usuarioC = UsuarioCorporativo.objects.get(usuario=user)
    codigo = usuarioC.codigo
    grupo= usuarioC.grupo
    chamados = Chamado.objects.filter(grupo=grupo).order_by("-id")    
    chamados_abertos = Chamado.objects.filter(active=True,grupo=grupo).order_by("-id")
    grupos= usuarioC.grupo.name

    filtro = ChamadoFilter()
    try:
        imageP = ImagePerfil.objects.get(nome= codigo.nome)
        if not imageP.image:
            imageP = ImagePerfil.objects.get(nome= "padrao")
    except:
        imageP = ImagePerfil.objects.get(nome= "padrao")          
      
    
       
    

 
    context = {
    'chamados': chamados,
    'chamados_abertos': chamados_abertos,
    'grupos': grupos,
    'usuarioC':usuarioC,
    'user' : user, 
    'grupo':grupo,
    'admin':admin, 
    'filtro': filtro,
    'imageP':imageP,    
            }  

    if request.method == 'POST' and 'salvar' in request.POST:
        
        nome = request.POST["name"]
        problema = request.POST["problem"]
        urgencia = request.POST["urgency"]
        if not nome:
                messages.error(
                    request, "Por favor preencha os campos ")
                return render(request, 'home/home.html', context)

        if not problema:
                messages.error(request, "Por favor escolha um assunto para tratar")
                return render(request, 'home/home.html', context)
        if not urgencia:
                messages.error(request, "Por favor escolha a urgencia do assunto")
                return render(request, 'home/home.html', context)        
        Chamado.objects.create(username=nome, problem=problema, urgency=urgencia)
        messages.success(request, 'Chamado criado com sucesso')



    return render(request, 'users/home.html', context) 
@login_required(login_url='/authentication/login')  
def presidente(request):
    user = request.user
    admin = Chamado.objects.all()        
    usuarioC = UsuarioCorporativo.objects.get(usuario=user)
    users = UsuarioCorporativo.objects.all()
    grupo= usuarioC.grupo
    chamados = Chamado.objects.filter(grupo=grupo).order_by("-id")    
    chamados_abertos = Chamado.objects.filter(active=True,grupo=grupo).order_by("-id")
    grupos= usuarioC.grupo.name
    codigo = usuarioC.codigo.nome
    imageP = ImagePerfil.objects.get(nome= codigo)
    if not imageP.image:
      imageP = ImagePerfil.objects.get(nome= "padrao")

    filtro = ChamadoFilter()
    context = {
    'chamados': chamados,
    'chamados_abertos': chamados_abertos,
    'grupos': grupos,
    'usuarioC':usuarioC,
    'user' : user,  
    'users' : users,  
    'grupo':grupo,
    'admin':admin, 
    'filtro': filtro,
    'imageP':imageP,    
            }  
    if request.method == 'GET':
        myFilter = ChamadoFilter(request.GET, queryset=users)
        users = myFilter.qs
               
        user = request.user
        usuarioC = UsuarioCorporativo.objects.get(usuario=user)
        grupo= usuarioC.grupo
        chamados_abertos = Chamado.objects.filter(active=True,grupo=grupo).order_by("id")
        grupos= usuarioC.grupo.name

        context = {
            'users':users,
            'chamados_abertos': chamados_abertos,
            'grupos': grupos,
            'usuarioC':usuarioC,
            'user' : user,  
            'grupo':grupo,
            'filtro': filtro,  
            'imageP':imageP,                   
                }  
    return render(request, 'users/presidente.html', context)     

MDATA = datetime.now().strftime('%d-%m-%Y')
def export_xlsx(model, filename, queryset, columns):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="%s"' % filename

    wb = xlwt.Workbook(encoding='utf-8')
    ws = wb.add_sheet(model)
    ws.col(0).width = 256 * 30
    ws.col(1).width = 256 * 30
    ws.col(2).width = 200 * 30
    ws.col(3).width = 200 * 30
    ws.col(4).width = 200 * 30
    
   
    row_num = 0

    font_style = xlwt.XFStyle()
    font_style.font.bold = True

    for col_num in range(len(columns)):
        ws.write(row_num, col_num, columns[col_num], font_style)

    default_style = xlwt.easyxf(num_format_str='dd/mm/yyyy')

    rows = queryset
    for row, rowdata in enumerate(rows):
        row_num += 1
        for col, val in enumerate(rowdata):
            ws.write(row_num, col, val, default_style)

    wb.save(response)
    return response
def exportar_colaboradores(request):
    MDATA = datetime.now().strftime('%Y-%m-%d')
    model = 'Users'
    filename = 'colaboradores_exportados.xls'
    _filename = filename.split('.')
    filename_final = f'{_filename[0]}_{MDATA}.{_filename[1]}'
    queryset = UsuarioCorporativo.objects.all().order_by("codigo__nome").values_list(
        'codigo__nome',
        'trabalho__cargo__cargo',
        'trabalho__departamento',
        'trabalho__dataadmissao',
        'codigo__datanacimento',
    )
    columns = ('Nome', 'Cargo','Departamento','Admissão','Data Nascimento')
    response = export_xlsx(model, filename_final, queryset,columns)
    return response

@login_required(login_url='/authentication/login')    
def problem(request): 
    user = request.user
    admin = Chamado.objects.all()        
    usuarioC = UsuarioCorporativo.objects.get(usuario=user)
    users = UsuarioCorporativo.objects.all()
    codigo = usuarioC.codigo
   
    grupo= usuarioC.grupo
    chamados = Chamado.objects.filter(grupo=grupo).order_by("-id")    
    chamados_abertos = Chamado.objects.filter(active=True,grupo=grupo).order_by("-id")
    grupos= usuarioC.grupo.name
    imageP = ImagePerfil.objects.get(nome= codigo.nome)
    if not imageP.image:
      imageP = ImagePerfil.objects.get(nome= "padrao")

    filtro = ChamadoFilter()
    context = {
            'chamados': chamados,
            'chamados_abertos': chamados_abertos,
            'grupos': grupos,
            'user' : user, 
            'users' : users,
            'usuarioC':usuarioC,
            
            'grupo':grupo,
            'filtro': filtro,  
            'imageP':imageP,                   
                }       
    if request.method == 'GET':
        myFilter = ChamadoFilter(request.GET, queryset=users)
        users = myFilter.qs 
        user = request.user
        usuarioC = UsuarioCorporativo.objects.get(usuario=user)
        grupo= usuarioC.grupo
        chamados_abertos = Chamado.objects.filter(active=True,grupo=grupo).order_by("id")
        grupo= usuarioC.grupo.name

        context = {
            'users':users,
            'chamados_abertos': chamados_abertos,
            'grupos': grupos,
            'usuarioC':usuarioC,
            
            'user' : user,  
            'grupo':grupo,
            'filtro': filtro,  
            'imageP':imageP,                   
                }            
    return render(request, 'users/problem.html', context)

@login_required(login_url='/authentication/login')    
def cargo(request):
    if request.method == 'POST' and 'cargo' in request.POST:
        cargo = request.POST["cargo"]
        ncbo = request.POST["ncbo"]
        Cargo.objects.create(cargo=cargo,ncbo=ncbo)
        messages.success(request,'Cadastrado com sucesso')
    return redirect ('add_usuarios.html')    


