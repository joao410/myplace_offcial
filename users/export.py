
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import load_workbook
from shutil import copyfile
import pandas as pd
from .models import UsuarioCorporativo 

user = UsuarioCorporativo.objects.all()




list_22 = []
df_22 = pd.DataFrame()
    
    
df_22 = df_22[['CM_COD_EST','CM_DES_PROD','CM_VAL_VRJ','CM_VAL_CMP']]
df_22['%'] = list_22





#####################################################################

template_file = 'C:\\Users\\arena\\projeto_python\\myplace\\TEMPLATE_RH.xlsx'
name = 'Relat_Colaboradores__' +  today.strftime("%d_%m_%Y") + '.xlsx' 
output_file = 'C:\\Users\\migue\\Documents\\Python_Desenvolvimento\\2021\\Compras\\' + name


copyfile(template_file, output_file)
wb = load_workbook(output_file)


#####################################################################


ws = wb['22']
ws['D2'] =today.strftime("%d/%m/%Y")



col_preco = ['C','D','E','F','G']


for i, r in zip(range(len(df_22)),dataframe_to_rows(df_22, index=False, header=False)):
        for col, item in zip(col_preco,r):
            ws[col+str(i+8)] =item


time.sleep(2)            
            
wb.save(output_file)  


print("Deu Certo Caraiii")
