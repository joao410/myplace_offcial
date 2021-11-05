from django.conf import settings
import pandas as pd
import fdb
from datetime import date,datetime, timedelta
import datetime as dt
import time
import os
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import load_workbook
from shutil import copyfile
import smtplib
import pyodbc


from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders


# def teste_excel():
    
#         today = date.today() + timedelta(days=1)
#         template_file = 'media\\excel_RH\\TEMPLATE_PROFISSIONAIS.xlsx'
#         name = 'Relat_Profissionais__' +  today.strftime("%d_%m_%Y") + '.xlsx' 
#         output_file = 'media\\excel_RH\\' + name
#         copyfile(template_file, output_file)
#         return output_file



def report():
        today = datetime.today()
        con = fdb.connect(dsn='arena/3050:C:\CICOM\MECAUTO\DB\CICOM.CDB', user='SYSDBA', password='masterkey')


        cur = con.cursor()

        #Tabela Vendas - Lembrar de abrir via Excel para checar os dados de dentro.
        SELECT = "SELECT * FROM TAB_CAD_FUN"
        cur.execute(SELECT)
        rows = cur.fetchall()
        prof = pd.DataFrame(rows)
        headers = [i[0] for i in cur.description]
        prof.columns = headers

        prof_2 = prof[['CM_COD_FUN','CM_NOM_FUN','CM_GENERO','CM_GRAU_INSTRU','CM_NAS_FUN','CM_CEP_FUN','CM_END_FUN','CM_NUM_END','CM_CID_FUN','CM_MAIL_FUN','CM_DAT_ADM','CM_REGIME_TRAB','CM_REGIME_PREV','CM_REGIME_JORN','CM_COD_CARGO','CM_COD_DEPTO','CM_DAT_DEM']]


        prof_2['CM_REGIME_TRAB'].replace({"1":"CLT","2":"RJU","3":"RJP"}, inplace=True)
        prof_2['CM_REGIME_PREV'].replace({"1":"RGPS","2":"RPPS","3":"RPPE"}, inplace=True)
        prof_2['CM_REGIME_JORN'].replace({"1":"Submetidos a Horário de Trabalho","2":"Atividade Externa Especificada","3":"Funções específicas"}, inplace=True)




        today = datetime.now()
        con = fdb.connect(dsn='arena/3050:C:\CICOM\MECAUTO\DB\CICOM.CDB', user='SYSDBA', password='masterkey')


        cur = con.cursor()

        #Tabela Vendas - Lembrar de abrir via Excel para checar os dados de dentro.
        SELECT = "SELECT * FROM TAB_CAD_CARGO"


        cur.execute(SELECT)
        rows = cur.fetchall()
        cargo = pd.DataFrame(rows)
        headers = [i[0] for i in cur.description]
        cargo.columns = headers




        lista_cargo=[]
        for item in prof_2['CM_COD_CARGO']:
            try:
                lista_cargo.append(cargo[cargo['CM_COD_CARGO']==int(item)]['CM_DESCRICAO'].item())
            except:
                lista_cargo.append('CARGO NAO DEFINIDO')
                
        prof_2['CARGO']=lista_cargo


        today = datetime.today()
        con = fdb.connect(dsn='arena/3050:C:\CICOM\MECAUTO\DB\CICOM.CDB', user='SYSDBA', password='masterkey')


        cur = con.cursor()

        #Tabela Vendas - Lembrar de abrir via Excel para checar os dados de dentro.
        SELECT = "SELECT * FROM TAB_CAD_DEPTO"


        cur.execute(SELECT)
        rows = cur.fetchall()
        depto = pd.DataFrame(rows)
        headers = [i[0] for i in cur.description]
        depto.columns = headers






        lista_dep=[]
        for item in prof_2['CM_COD_DEPTO']:
            try:
                lista_dep.append(depto[depto['CM_COD_DEPTO']==int(item)]['CM_DESCRICAO'].item())
            except:
                lista_dep.append('DEPT NAO DEFINIDO')
                
        prof_2['DEPT']=lista_dep




        prof_2.drop('CM_COD_CARGO',  axis='columns', inplace=True)
        prof_2.drop('CM_COD_DEPTO',  axis='columns', inplace=True)


        template_file = 'media\\excel_RH\\TEMPLATE_PROFISSIONAIS.xlsx'
        name = 'Relat_Profissionais__' +  today.strftime("%d_%m_%Y_%H_%M") + '.xlsx' 
        output_file =  settings.MEDIA_ROOT +  '\\models_rh\\' + name
      

        copyfile(template_file, output_file)
        wb = load_workbook(output_file)

        ws = wb['Profissionais']
        ws['D2'] =today.strftime("%d/%m/%Y-%H:%M:%S.%f")



        col_preco = ['C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S']


        for i, r in zip(range(len(prof_2)),dataframe_to_rows(prof_2, index=False, header=False)):
                for col, item in zip(col_preco,r):
                    ws[col+str(i+8)] =item



            
                    
        wb.save(output_file)
       

        
      
  
def atacado():
    today = date.today()
    con = fdb.connect(dsn='arena/3050:C:\CICOM\MECAUTO\DB\CICOM.CDB', user='SYSDBA', password='masterkey')


    cur = con.cursor()

    #Tabela Vendas - Lembrar de abrir via Excel para checar os dados de dentro.
    SELECT = "SELECT CM_COD_EST,CM_DES_PROD,CM_VAL_CMP, CM_VAL_VRJ ,CM_VAL_VND  FROM TAB_CAD_PROD WHERE CM_DES_TAQUE='N'"


    cur.execute(SELECT)
    rows = cur.fetchall()
    prod = pd.DataFrame(rows)
    headers = [i[0] for i in cur.description]
    prod.columns = headers

    time.sleep(2)

    #######################################################################################


    list_22 = []
    df_22 = pd.DataFrame()

    list_60 = []
    df_60 = pd.DataFrame()

    for i in range(len(prod)):
        try:
            perc_22 = round((prod['CM_VAL_VRJ'].iloc[i]/prod['CM_VAL_CMP'].iloc[i] - 1)*100,2)
            perc_60 = round((prod['CM_VAL_VND'].iloc[i]/prod['CM_VAL_CMP'].iloc[i] - 1)*100,2)

            if perc_22 < 22:
                list_22.append(perc_22)
                df_22 = df_22.append(prod.iloc[i])

            if perc_60 <60:
                list_60.append(perc_60)
                df_60 = df_60.append(prod.iloc[i])
        except:
            print(i)
        
        
    time.sleep(4)
        
        
    df_22 = df_22[['CM_COD_EST','CM_DES_PROD','CM_VAL_VRJ','CM_VAL_CMP']]
    df_22['%'] = list_22

    df_60 = df_60[['CM_COD_EST','CM_DES_PROD','CM_VAL_VND','CM_VAL_CMP']]    
    df_60['%'] = list_60




    #####################################################################

    template_file = 'C:\\Users\\arena\\myplace\\media\\Comercial\\TEMPLATE_ATACADO.xlsx'
    name = 'Relat_Precos__' +  today.strftime("%d_%m_%Y") + '.xlsx' 
    output_file = 'C:\\Users\\arena\\myplace\\media\\Comercial\\' + name


    copyfile(template_file, output_file)
    wb = load_workbook(output_file)


    #####################################################################


    ws = wb['22']
    ws['D2'] =today.strftime("%d/%m/%Y")



    col_preco = ['C','D','E','F','G']


    for i, r in zip(range(len(df_22)),dataframe_to_rows(df_22, index=False, header=False)):
            for col, item in zip(col_preco,r):
                ws[col+str(i+8)] =item

    ws = wb['60']
    ws['D2'] =today.strftime("%d/%m/%Y")


    for i, r in zip(range(len(df_60)),dataframe_to_rows(df_60, index=False, header=False)):
            for col, item in zip(col_preco,r):
                ws[col+str(i+8)] =item

    time.sleep(2)            
                
    wb.save(output_file)  


    print("20_60 - OK")
    time.sleep(5)


def estoque():
    today = date.today()
    data = today -  timedelta(days=3)
    con = fdb.connect(dsn='arena/3050:C:\CICOM\MECAUTO\DB\CICOM.CDB', user='SYSDBA', password='masterkey')

    cur = con.cursor()

    #Tabela Vendas - Lembrar de abrir via Excel para checar os dados de dentro.
    SELECT = "SELECT CM_COD_EST,CM_DES_PROD,CM_COD_FAB,CM_COD_FOR,CM_QTD_EST,CM_MIN_EST,CM_MAX_EST , CM_CUSTO_MEDIO FROM TAB_CAD_PROD "


    cur.execute(SELECT)
    rows = cur.fetchall()
    prod = pd.DataFrame(rows)
    headers = [i[0] for i in cur.description]
    prod.columns = headers



    prod_1 = prod[prod['CM_QTD_EST']>=0]
    prod_2 = prod_1[prod_1['CM_MIN_EST']>2]
    prod_f = prod_2[prod_2['CM_QTD_EST']<prod_2['CM_MIN_EST']].sort_values(by='CM_MIN_EST', ascending=False)
    prod_neg = prod[prod['CM_QTD_EST']<0]



    #####################################################################

   
    template_file = 'C:\\Users\\arena\\myplace\\media\\Comercial\\\TEMPLATE_FALTANTE.xlsx'
    name = 'Relat_Faltantes_Estoque__' +  today.strftime("%d_%m_%Y") + '.xlsx' 
    output_file = 'C:\\Users\\arena\\myplace\\media\\Comercial\\' + name


    copyfile(template_file, output_file)
    wb = load_workbook(output_file)


    #####################################################################


    ws = wb['FALTA']
    ws['D2'] =today.strftime("%d/%m/%Y")



    col_preco = ['C','D','E','F','G','H','I','J']

    for i, r in zip(range(len(prod_f)),dataframe_to_rows(prod_f, index=False, header=False)):
            for col, item in zip(col_preco,r):
                ws[col+str(i+8)] =item
            
                
    wb.save(output_file)  


    ws = wb['NEGATIVO']
    ws['D2'] =today.strftime("%d/%m/%Y")

    col_preco = ['C','D','E','F','G','H','I','J']

    for i, r in zip(range(len(prod_neg)),dataframe_to_rows(prod_neg, index=False, header=False)):
            for col, item in zip(col_preco,r):
                ws[col+str(i+8)] =item
                
                
    wb.save(output_file)  


    print("ESTOQUE FALTANTE - OK")
    time.sleep(5)

def compras():
    week_day = date.today().weekday()

    if week_day != 0:
            
        today = date.today()# - timedelta(days=1)
        today_d = today.day 
        today_m = today.month
        today_Y = today.year

        server = 'ARENA' 
        database = 'CADASTRO' 
        username = 'tigenios' 
        password = '0567senh@' 
        cnxn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER='+server+';DATABASE='+database+';UID='+username+';PWD='+ password)
        cursor = cnxn.cursor()
        SQL = f"SELECT * FROM RELATORIO_COMPRAS WHERE datepart(day,DATE)='{today_d}' AND datepart(month,DATE)='{today_m}' AND datepart(year,DATE)='{today_Y}'"

        cursor.execute(SQL)
        rows = cursor.fetchall()
        headers = [i[0] for i in cursor.description]

        df = pd.DataFrame.from_records(rows, columns =headers)



        ip_file = 'C:\\Users\\arena\\myplace\\media\\IP_USER.xlsx'

        wb = load_workbook(ip_file)
        ws = wb['Sheet1']

        dic_ip = {}
        for i in range(len(ws['A'])):
            dic_ip[ws['A'+str(i+2)].value] = ws['B'+str(i+2)].value

        for i in range(len(df)):
            ip_fim = df['USER_CHANGE'].iloc[i].find('/')
            try:
                df['USER_CHANGE'].iloc[i] = dic_ip[df['USER_CHANGE'].iloc[i][0:ip_fim]]
            except:
                pass

        df_change = df[df['TYPE']=='CHANGE'][['PRODUCT','COD_MEC','VALUE_BEFORE_BUY','VALUE_BUY','VALUE_BEFORE_SELL','VALUE_SELL','USER_CHANGE','DATE']]
        df_new = df[df['TYPE']=='NEW'][['PRODUCT','COD_MEC','VALUE_BUY','VALUE_SELL','USER_CHANGE','DATE']]



        template_file = 'C:\\Users\\arena\\myplace\\media\\Comercial\\Relat_Compras.xlsx' 

        name = 'Relat_Compras_' +  today.strftime("%d_%m_%Y") + '.xlsx' 

        output_file = 'C:\\Users\\arena\\myplace\\media\\Comercial\\' + name


        col_new = ['A','B','C','D','E','F']
        col_change = ['A','B','C','D','E','F','G','H']

        copyfile(template_file, output_file)
        wb = load_workbook(output_file)

        ### NEW #####
        ws = wb['Novo']
        for i, r in zip(range(len(df_new)),dataframe_to_rows(df_new, index=False, header=False)):
            for col, item in zip(col_new,r):
                ws[col+str(i+2)] =item
        wb.save(output_file)    

        ### CHANGE  ####
        ws = wb['Mudou']
        for i, r in zip(range(len(df_change)),dataframe_to_rows(df_change, index=False, header=False)):
            for col, item in zip(col_change,r):
                ws[col+str(i+2)] =item
        wb.save(output_file) 


        ####################### EMAIL ##########################


    #     try:
    #         fromaddr = "tigenioshop@gmail.com"
    #         toaddr = ['andre@arenavidros.com.br','elton@arenavidros.com.br','deise@arenavidros.com.br','antonio@arenavidros.com.br']
    #         msg = MIMEMultipart()

    #         msg['From'] = fromaddr 
    #         msg['To'] = ", ".join(toaddr)
    #         msg['Subject'] = "Relatório de Compras " + today.strftime("%d/%m/%Y") 

    #         body = "\nSegue em anexo o Relatório de Compras do dia " + today.strftime("%d/%m/%Y") + ",\nna aba 'Novo' constam os produtos adicionados, na aba 'Mudou' constam os produtos modificados.\nAtt,\nAndré Miguel G. Teiga"

    #         msg.attach(MIMEText(body, 'plain'))

    #         filename = output_file

    #         attachment = open(filename,'rb')


    #         part = MIMEBase('application', 'octet-stream')
    #         part.set_payload((attachment).read())
    #         encoders.encode_base64(part)
    #         part.add_header('Content-Disposition', "attachment; filename= %s" % name)

    #         msg.attach(part)

    #         attachment.close()

    #         server = smtplib.SMTP('smtp.gmail.com', 587)
    #         server.starttls()
    #         server.login(fromaddr, "0567Senh@")
    #         text = msg.as_string()
    #         server.sendmail(fromaddr, toaddr, text)
    #         server.quit()
    #         print('\nEmail enviado com sucesso!')
    #     except:
    #         print("\nErro ao enviar email")

        print("RELAT COMPRAS - OK")
        time.sleep(5)

    else:
        print('Boa Segunda Feira')
        time.sleep(5)

        